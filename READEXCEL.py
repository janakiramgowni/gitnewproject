import openpyxl
path="C:\Book.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active  # if there is only one sheet
#sheet=workbook.get_sheet_by_name("sheet1")
rows=sheet.max_row  # will get no of rows
cols=sheet.max_column
print(rows) #7
print(cols) #3


for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value,end="         ")
    print()
